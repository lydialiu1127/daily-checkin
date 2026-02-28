import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "每日作息安排表"

# 颜色定义
purple = "6C63FF"
light_purple = "F0EEFF"
white = "FFFFFF"
gray_text = "8E8BA3"
green = "4CAF50"

# 边框
thin_border = Border(
    left=Side(style='thin', color='E8E6F0'),
    right=Side(style='thin', color='E8E6F0'),
    top=Side(style='thin', color='E8E6F0'),
    bottom=Side(style='thin', color='E8E6F0')
)

# 列宽
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 8

# ===== 标题行 =====
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = "✨ 我的美好一天 · 每日作息安排表 ✨"
title_cell.font = Font(name='PingFang SC', size=18, bold=True, color=purple)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = PatternFill(start_color=white, end_color=white, fill_type='solid')
ws.row_dimensions[1].height = 45

# 副标题
ws.merge_cells('A2:E2')
sub_cell = ws['A2']
sub_cell.value = "自律即自由 · 日期：____________"
sub_cell.font = Font(name='PingFang SC', size=11, color=gray_text)
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 28

# ===== 表头 =====
headers = ['', '时间', '事项', '备注', '✓']
header_fill = PatternFill(start_color=purple, end_color=purple, fill_type='solid')
header_font = Font(name='PingFang SC', size=11, bold=True, color=white)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
ws.row_dimensions[3].height = 30

# ===== 数据 =====
schedule = [
    # (类型, 图标, 时间, 事项, 备注)  类型: 'section' 或 'item'
    ('section', '', '', '🌅 早间 · 唤醒身体', ''),
    ('item', '⏰', '6:20', '起床', '新的一天开始啦'),
    ('item', '💪', '6:20 – 6:25', '鼓肚子 5 分钟', '唤醒身体'),
    ('item', '🩸', '6:25 – 6:30', '气血操 5 分钟', '促进气血循环'),
    ('item', '🧘', '6:30 – 7:30', '瑜伽 60 分钟', '舒展身心，保持健康'),
    ('item', '🥋', '7:30 – 7:45', '八段锦 15 分钟', '传统养生功法'),
    ('section', '', '', '💄 上午 · 出门准备', ''),
    ('item', '💄', '7:45 – 8:20', '洗漱 · 化妆 · 穿搭', '同时听播客'),
    ('item', '🚗', '8:20 – 9:00', '上班 + 公司早餐', '上班路上听帆书'),
    ('section', '', '', '💻 工作 · 专注高效', ''),
    ('item', '📋', '9:00 – 9:50', '列出今日重点工作', '查看邮件，理清思路'),
    ('item', '🔥', '9:50', '切换工作状态', '进入专注模式'),
    ('item', '💻', '10:00 – 12:00', '专注高效工作（上午）', '心无旁骛，全力以赴'),
    ('section', '', '', '🌿 午间 · 自由充电', ''),
    ('item', '🌿', '12:00 – 14:00', '午间自由安排', '走路听书/美容院/午休冥想'),
    ('section', '', '', '💻 下午 · 高效产出', ''),
    ('item', '💻', '14:00 – 18:00', '专注高效工作（下午）', '保持节奏，高效产出'),
    ('section', '', '', '🌙 晚间 · 放松成长', ''),
    ('item', '🍽️', '18:00 – 20:00', '下班 · 晚餐 · 饭后放松', '饭后站立放松'),
    ('item', '📚', '20:00 – 21:00', '学习一小时', '理财/化妆/穿搭等'),
    ('item', '🦶', '21:00 – 21:30', '泡脚', '敷面膜/听古典音乐/看微博'),
    ('item', '📝', '21:30 – 22:00', '总结日记 · 准备明日穿搭', '回顾今天，规划明天'),
    ('item', '🌙', '22:00 – 23:00', '睡前仪式', '五动作 · 冥想 · 阅读'),
]

section_fill = PatternFill(start_color=light_purple, end_color=light_purple, fill_type='solid')
section_font = Font(name='PingFang SC', size=11, bold=True, color=purple)
item_font = Font(name='PingFang SC', size=11, color='2D2B55')
time_font = Font(name='PingFang SC', size=11, bold=True, color=purple)
note_font = Font(name='PingFang SC', size=10, color=gray_text)
icon_font = Font(name='PingFang SC', size=14)
even_fill = PatternFill(start_color='FAFAFF', end_color='FAFAFF', fill_type='solid')
white_fill = PatternFill(start_color=white, end_color=white, fill_type='solid')

row = 4
item_count = 0
for entry in schedule:
    typ, icon, time_str, task, note = entry
    
    if typ == 'section':
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value=task)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = thin_border
        ws.row_dimensions[row].height = 28
    else:
        item_count += 1
        fill = even_fill if item_count % 2 == 0 else white_fill
        
        # 图标
        c1 = ws.cell(row=row, column=1, value=icon)
        c1.font = icon_font
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.fill = fill
        c1.border = thin_border
        
        # 时间
        c2 = ws.cell(row=row, column=2, value=time_str)
        c2.font = time_font
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.fill = fill
        c2.border = thin_border
        
        # 事项
        c3 = ws.cell(row=row, column=3, value=task)
        c3.font = item_font
        c3.alignment = Alignment(horizontal='left', vertical='center')
        c3.fill = fill
        c3.border = thin_border
        
        # 备注
        c4 = ws.cell(row=row, column=4, value=note)
        c4.font = note_font
        c4.alignment = Alignment(horizontal='left', vertical='center')
        c4.fill = fill
        c4.border = thin_border
        
        # 打勾框
        c5 = ws.cell(row=row, column=5, value='☐')
        c5.font = Font(name='PingFang SC', size=14, color='C8C6D6')
        c5.alignment = Alignment(horizontal='center', vertical='center')
        c5.fill = fill
        c5.border = thin_border
        
        ws.row_dimensions[row].height = 32
    
    row += 1

# ===== 底部 =====
row += 1
ws.merge_cells(f'A{row}:E{row}')
footer = ws.cell(row=row, column=1, value='✨ 自律的每一天，都是在靠近更好的自己 ✨')
footer.font = Font(name='PingFang SC', size=12, bold=True, color=purple)
footer.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 35

row += 1
ws.merge_cells(f'A{row}:E{row}')
footer2 = ws.cell(row=row, column=1, value='完成率：____ / 17')
footer2.font = Font(name='PingFang SC', size=11, color=gray_text)
footer2.alignment = Alignment(horizontal='center', vertical='center')

# 打印设置
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1

# 保存
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '每日作息安排表.xlsx')
wb.save(output_path)
print(f'已生成：{output_path}')
